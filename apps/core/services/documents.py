"""
Server-side document generation:
  - Payroll/payslip PDFs (reportlab), optionally password-protected (pikepdf)
  - Color-styled Excel payroll reports (openpyxl) — client-requested styling
  - SHA-256 fingerprinting for tamper evidence / audit trail

Payroll calculation never leaves this server; DocuSeal receives only the
rendered PDF for signature, not the underlying figures' logic.
"""
import hashlib
import io
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer, Table,
                                TableStyle)
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet

BRAND = colors.HexColor('#1d4ed8')
BRAND_LIGHT = colors.HexColor('#dbeafe')


def sha256_of(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _money(v):
    return f"KES {Decimal(v or 0):,.2f}"


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

def generate_payroll_pdf(run, records, *, company_name='', triggered_by='',
                         watermark_status=None) -> bytes:
    """
    Render a payroll run summary PDF: header with audit info, one row per
    employee with gross/deductions/net, totals footer.
    `records` are PayrollRecord instances (or dicts with the same keys).
    """
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=18 * mm, bottomMargin=18 * mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('t', parent=styles['Title'], textColor=BRAND)
    small = ParagraphStyle('s', parent=styles['Normal'], fontSize=8, textColor=colors.grey)

    story = [
        Paragraph(f"{company_name or 'Company'} — Payroll {run.period_display}", title_style),
        Paragraph(
            f"Run ID: {run.id} &nbsp;&nbsp; Status: {run.status}"
            f"{' • ' + watermark_status if watermark_status else ''}", small),
        Paragraph(f"Triggered by: {triggered_by or run.run_by} • Generated server-side; "
                  f"figures are final at signature time.", small),
        Spacer(1, 6 * mm),
    ]

    header = ['#', 'Employee', 'Job title', 'Gross', 'PAYE', 'NSSF', 'NHIF/SHIF',
              'HELB', 'Other', 'Net']
    rows = [header]
    for i, rec in enumerate(records, 1):
        emp = getattr(rec, 'employee', None)
        rows.append([
            str(i),
            getattr(emp, 'employee_number', '') if emp else rec.get('employee_number', ''),
            getattr(emp, 'job_title', '') if emp else rec.get('job_title', ''),
            _money(rec.gross_salary if emp else rec.get('gross_salary')),
            _money(rec.paye if emp else rec.get('paye')),
            _money(rec.nssf if emp else rec.get('nssf')),
            _money(rec.nhif if emp else rec.get('nhif')),
            _money(rec.helb if emp else rec.get('helb')),
            _money(rec.other_deductions if emp else rec.get('other_deductions')),
            _money(rec.net_salary if emp else rec.get('net_salary')),
        ])
    rows.append(['', 'TOTALS', '', _money(run.total_gross), '', '', '', '',
                 _money(run.total_deductions), _money(run.total_net)])

    table = Table(rows, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), BRAND),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cbd5e1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, BRAND_LIGHT]),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#1e293b')),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
        ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),
    ]))
    story.append(table)
    doc.build(story)
    return buf.getvalue()


def generate_payslip_pdf(record, *, company_name='', period='', allowances=None,
                         extra_deductions=None) -> bytes:
    """Single-employee payslip with dynamic allowance/deduction lines."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20 * mm)
    styles = getSampleStyleSheet()
    emp = record.employee
    story = [
        Paragraph(f"{company_name} — Payslip {period}", styles['Title']),
        Paragraph(f"{emp.employee_number} • {emp.job_title} • {emp.department or ''}",
                  styles['Normal']),
        Spacer(1, 6 * mm),
    ]
    rows = [['Item', 'Amount'], ['Gross salary', _money(record.gross_salary)]]
    for a in allowances or []:
        rows.append([f"Allowance: {a['name']}", _money(a['amount'])])
    rows += [
        ['PAYE', f"-{_money(record.paye)}"],
        ['NSSF', f"-{_money(record.nssf)}"],
        ['NHIF/SHIF', f"-{_money(record.nhif)}"],
        ['HELB', f"-{_money(record.helb)}"],
    ]
    for d in extra_deductions or []:
        rows.append([f"Deduction: {d['name']}", f"-{_money(d['amount'])}"])
    rows.append(['NET PAY', _money(record.net_salary)])
    t = Table(rows, colWidths=[100 * mm, 60 * mm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), BRAND),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cbd5e1')),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#16a34a')),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
    ]))
    story.append(t)
    doc.build(story)
    return buf.getvalue()


def password_protect_pdf(pdf_bytes: bytes, user_password: str,
                         owner_password: str | None = None) -> bytes:
    """AES-256 encrypt a PDF (pikepdf/qpdf)."""
    import pikepdf
    src = io.BytesIO(pdf_bytes)
    out = io.BytesIO()
    with pikepdf.open(src) as pdf:
        pdf.save(out, encryption=pikepdf.Encryption(
            user=user_password,
            owner=owner_password or user_password,
            R=6,
            allow=pikepdf.Permissions(extract=False, modify_annotation=False,
                                      modify_assembly=False, modify_form=False,
                                      modify_other=False),
        ))
    return out.getvalue()


# ---------------------------------------------------------------------------
# Excel (color-styled, per client request)
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill('solid', fgColor='1D4ED8')
ALT_FILL = PatternFill('solid', fgColor='DBEAFE')
TOTAL_FILL = PatternFill('solid', fgColor='1E293B')
NET_FILL = PatternFill('solid', fgColor='16A34A')
THIN = Border(*[Side(style='thin', color='CBD5E1')] * 4)


def generate_payroll_excel(run, records, *, company_name='') -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = f"Payroll {run.period_month}-{run.period_year}"

    ws.merge_cells('A1:J1')
    c = ws['A1']
    c.value = f"{company_name or 'Company'} — Payroll {run.period_display} (status: {run.status})"
    c.font = Font(bold=True, size=14, color='1D4ED8')

    headers = ['#', 'Employee No', 'Job Title', 'Gross', 'PAYE', 'NSSF',
               'NHIF/SHIF', 'HELB', 'Other Deductions', 'Net Pay']
    ws.append([])
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN

    row_idx = 4
    for i, rec in enumerate(records, 1):
        emp = rec.employee
        values = [i, emp.employee_number, emp.job_title,
                  float(rec.gross_salary), float(rec.paye), float(rec.nssf),
                  float(rec.nhif), float(rec.helb), float(rec.other_deductions),
                  float(rec.net_salary)]
        ws.append(values)
        for col in range(1, len(values) + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = THIN
            if i % 2 == 0:
                cell.fill = ALT_FILL
            if col >= 4:
                cell.number_format = '#,##0.00'
        ws.cell(row=row_idx, column=10).font = Font(bold=True, color='16A34A')
        row_idx += 1

    totals = ['', 'TOTALS', '', float(run.total_gross), '', '', '', '',
              float(run.total_deductions), float(run.total_net)]
    ws.append(totals)
    for col in range(1, 11):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = TOTAL_FILL
        cell.font = Font(bold=True, color='FFFFFF')
        cell.border = THIN
        if col >= 4:
            cell.number_format = '#,##0.00'
    ws.cell(row=row_idx, column=10).fill = NET_FILL

    widths = [5, 16, 24, 14, 12, 12, 12, 12, 16, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A4'

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
