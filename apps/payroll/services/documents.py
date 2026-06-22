"""
Server-side document generation:
  - Payroll/payslip PDFs (reportlab), optionally password-protected (pikepdf)
  - Color-styled Excel payroll reports (openpyxl) — client-requested styling
  - SHA-256 fingerprinting for tamper evidence / audit trail

Payroll calculation never leaves this server; DocuSeal receives only the
rendered PDF for signature, not the underlying figures' logic.
"""
import hashlib
import io
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer, Table,
                                TableStyle)
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet

BRAND = colors.HexColor('#F47920')       # Sheer Logic orange
BRAND_LIGHT = colors.HexColor('#FDE8D0')  # light orange tint


def sha256_of(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _money(v):
    return f"KES {Decimal(v or 0):,.2f}"


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

def generate_payroll_pdf(run, records, *, company_name='', triggered_by='',
                         watermark_status=None) -> bytes:
    """
    Render a payroll run summary PDF: header with audit info, one row per
    employee with gross/deductions/net, totals footer.
    `records` are PayrollRecord instances (or dicts with the same keys).
    """
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=18 * mm, bottomMargin=18 * mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('t', parent=styles['Title'], textColor=BRAND)
    small = ParagraphStyle('s', parent=styles['Normal'], fontSize=8, textColor=colors.grey)  # noqa

    story = [
        Paragraph(f"{company_name or 'Company'} — Payroll {run.period_display}", title_style),
        Paragraph(
            f"Run ID: {run.id} &nbsp;&nbsp; Status: {run.status}"
            f"{' • ' + watermark_status if watermark_status else ''}", small),
        Paragraph(f"Triggered by: {triggered_by or run.run_by} • Generated server-side; "
                  f"figures are final at signature time.", small),
        Spacer(1, 6 * mm),
    ]

    header = ['#', 'Employee', 'Job title', 'Gross', 'PAYE', 'NSSF', 'NHIF/SHIF',
              'HELB', 'Other', 'Net']
    rows = [header]
    for i, rec in enumerate(records, 1):
        emp = getattr(rec, 'employee', None)
        rows.append([
            str(i),
            getattr(emp, 'employee_number', '') if emp else rec.get('employee_number', ''),
            getattr(emp, 'job_title', '') if emp else rec.get('job_title', ''),
            _money(rec.gross_salary if emp else rec.get('gross_salary')),
            _money(rec.paye if emp else rec.get('paye')),
            _money(rec.nssf if emp else rec.get('nssf')),
            _money(rec.nhif if emp else rec.get('nhif')),
            _money(rec.helb if emp else rec.get('helb')),
            _money(rec.other_deductions if emp else rec.get('other_deductions')),
            _money(rec.net_salary if emp else rec.get('net_salary')),
        ])
    rows.append(['', 'TOTALS', '', _money(run.total_gross), '', '', '', '',
                 _money(run.total_deductions), _money(run.total_net)])

    table = Table(rows, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), BRAND),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cbd5e1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, BRAND_LIGHT]),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#1e293b')),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
        ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),
    ]))
    story.append(table)
    doc.build(story)
    return buf.getvalue()


def generate_payslip_pdf(record, *, company_name='', period='', allowances=None,
                         extra_deductions=None) -> bytes:
    """Single-employee payslip with dynamic allowance/deduction lines."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20 * mm)
    styles = getSampleStyleSheet()
    emp = record.employee
    story = [
        Paragraph(f"{company_name} — Payslip {period}", styles['Title']),
        Paragraph(f"{emp.employee_number} • {emp.job_title} • {emp.department or ''}",
                  styles['Normal']),
        Spacer(1, 6 * mm),
    ]
    rows = [['Item', 'Amount'], ['Gross salary', _money(record.gross_salary)]]
    for a in allowances or []:
        rows.append([f"Allowance: {a['name']}", _money(a['amount'])])
    rows += [
        ['PAYE', f"-{_money(record.paye)}"],
        ['NSSF', f"-{_money(record.nssf)}"],
        ['NHIF/SHIF', f"-{_money(record.nhif)}"],
        ['HELB', f"-{_money(record.helb)}"],
    ]
    for d in extra_deductions or []:
        rows.append([f"Deduction: {d['name']}", f"-{_money(d['amount'])}"])
    rows.append(['NET PAY', _money(record.net_salary)])
    t = Table(rows, colWidths=[100 * mm, 60 * mm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), BRAND),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cbd5e1')),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#16a34a')),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
    ]))
    story.append(t)
    doc.build(story)
    return buf.getvalue()


def password_protect_pdf(pdf_bytes: bytes, user_password: str,
                         owner_password: str | None = None) -> bytes:
    """AES-256 encrypt a PDF (pikepdf/qpdf)."""
    import pikepdf
    src = io.BytesIO(pdf_bytes)
    out = io.BytesIO()
    with pikepdf.open(src) as pdf:
        pdf.save(out, encryption=pikepdf.Encryption(
            user=user_password,
            owner=owner_password or user_password,
            R=6,
            allow=pikepdf.Permissions(extract=False, modify_annotation=False,
                                      modify_assembly=False, modify_form=False,
                                      modify_other=False),
        ))
    return out.getvalue()


# ---------------------------------------------------------------------------
# Excel (2-sheet: Summary + Payroll Detail)
# ---------------------------------------------------------------------------

BRAND_HEX = 'F47920'          # Sheer Logic orange
BRAND_LIGHT_HEX = 'FDE8D0'
HEADER_FILL = PatternFill('solid', fgColor=BRAND_HEX)
ALT_FILL = PatternFill('solid', fgColor=BRAND_LIGHT_HEX)
TOTAL_FILL = PatternFill('solid', fgColor='1E293B')
NET_FILL = PatternFill('solid', fgColor='16A34A')
KPI_FILLS = {
    'employees': PatternFill('solid', fgColor='F47920'),
    'gross':     PatternFill('solid', fgColor='1E6B3C'),
    'deductions':PatternFill('solid', fgColor='7C3AED'),
    'net':       PatternFill('solid', fgColor='16A34A'),
}
THIN = Border(*[Side(style='thin', color='CBD5E1')] * 4)
NO_BORDER = Border()


def _kpi_box(ws, col, label, value, fill):
    """Write a 2-row KPI box: label on row 4, value on row 5."""
    lbl = ws.cell(row=4, column=col, value=label)
    lbl.fill = fill
    lbl.font = Font(bold=True, color='FFFFFF', size=10)
    lbl.alignment = Alignment(horizontal='center', vertical='center')
    lbl.border = NO_BORDER

    val = ws.cell(row=5, column=col, value=value)
    val.fill = fill
    val.font = Font(bold=True, color='FFFFFF', size=14)
    val.alignment = Alignment(horizontal='center', vertical='center')
    val.border = NO_BORDER
    if isinstance(value, float):
        val.number_format = '#,##0.00'


def _build_summary_sheet(ws, run, records, company_name):
    ws.sheet_view.showGridLines = False

    # --- Banner (rows 1-2) ---
    ws.merge_cells('A1:H1')
    banner = ws['A1']
    banner.value = company_name or 'Company'
    banner.font = Font(bold=True, size=18, color=BRAND_HEX)
    banner.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 32

    ws.merge_cells('A2:H2')
    sub = ws['A2']
    sub.value = f"Payroll Summary — {run.period_display}  |  Status: {run.status.upper()}"
    sub.font = Font(size=11, color='555555')
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 12  # spacer

    # --- KPI boxes (rows 4-5, cols A-D) ---
    total_deductions = float(run.total_deductions or 0)
    total_gross = float(run.total_gross or 0)
    pct = f"{(total_deductions / total_gross * 100):.1f}%" if total_gross else '0%'

    _kpi_box(ws, 1, 'Employees', len(records), KPI_FILLS['employees'])
    _kpi_box(ws, 3, 'Total Gross (KES)', float(run.total_gross or 0), KPI_FILLS['gross'])
    _kpi_box(ws, 5, 'Total Deductions (KES)', total_deductions, KPI_FILLS['deductions'])
    _kpi_box(ws, 7, 'Net Pay (KES)', float(run.total_net or 0), KPI_FILLS['net'])
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 28
    ws.row_dimensions[6].height = 14  # spacer

    # --- Deductions breakdown table (from row 7) ---
    breakdown_headers = ['Deduction', 'Total (KES)', '% of Gross']
    for ci, h in enumerate(breakdown_headers, 1):
        cell = ws.cell(row=7, column=ci, value=h)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN

    deductions = [
        ('PAYE', sum(float(r.paye or 0) for r in records)),
        ('NSSF', sum(float(r.nssf or 0) for r in records)),
        ('NHIF / SHIF', sum(float(r.nhif or 0) for r in records)),
        ('HELB', sum(float(r.helb or 0) for r in records)),
        ('Other', sum(float(r.other_deductions or 0) for r in records)),
    ]
    for ri, (name, amount) in enumerate(deductions, 8):
        ws.cell(row=ri, column=1, value=name).border = THIN
        amt_cell = ws.cell(row=ri, column=2, value=amount)
        amt_cell.number_format = '#,##0.00'
        amt_cell.border = THIN
        pct_val = f"{(amount / total_gross * 100):.1f}%" if total_gross else '0%'
        ws.cell(row=ri, column=3, value=pct_val).border = THIN
        if ri % 2 == 1:
            for ci in range(1, 4):
                ws.cell(row=ri, column=ci).fill = ALT_FILL

    # Totals row
    tr = 8 + len(deductions)
    ws.cell(row=tr, column=1, value='TOTAL DEDUCTIONS').font = Font(bold=True, color='FFFFFF')
    ws.cell(row=tr, column=1).fill = TOTAL_FILL
    tc = ws.cell(row=tr, column=2, value=total_deductions)
    tc.number_format = '#,##0.00'
    tc.font = Font(bold=True, color='FFFFFF')
    tc.fill = TOTAL_FILL
    ws.cell(row=tr, column=3, value=pct).font = Font(bold=True, color='FFFFFF')
    ws.cell(row=tr, column=3).fill = TOTAL_FILL

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 14
    for col in ('D', 'E', 'F', 'G', 'H'):
        ws.column_dimensions[col].width = 18


def _build_detail_sheet(ws, run, records, company_name):
    ws.sheet_view.showGridLines = False

    # Branded header rows
    ws.merge_cells('A1:L1')
    h1 = ws['A1']
    h1.value = company_name or 'Company'
    h1.font = Font(bold=True, size=16, color=BRAND_HEX)
    h1.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:L2')
    h2 = ws['A2']
    h2.value = f"Payroll Detail — {run.period_display}  |  Status: {run.status.upper()}"
    h2.font = Font(size=10, color='555555')
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 10  # spacer

    headers = ['#', 'Employee No', 'Full Name', 'Department', 'Job Title',
               'Gross', 'PAYE', 'NSSF', 'NHIF/SHIF', 'HELB',
               'Other Deductions', 'Net Pay']
    ws.append([])  # row 4 = headers (append adds to next empty row after merges)
    # openpyxl append goes to next row; we need row 4
    # Re-write header at row 4 explicitly
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN
    ws.row_dimensions[4].height = 18

    row_idx = 5
    for i, rec in enumerate(records, 1):
        emp = rec.employee
        full_name = f"{getattr(emp, 'first_name', '')} {getattr(emp, 'last_name', '')}".strip()
        values = [
            i,
            emp.employee_number,
            full_name or emp.employee_number,
            getattr(emp, 'department', '') or '',
            emp.job_title or '',
            float(rec.gross_salary or 0),
            float(rec.paye or 0),
            float(rec.nssf or 0),
            float(rec.nhif or 0),
            float(rec.helb or 0),
            float(rec.other_deductions or 0),
            float(rec.net_salary or 0),
        ]
        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=ci, value=val)
            cell.border = THIN
            if i % 2 == 0:
                cell.fill = ALT_FILL
            if ci >= 6:
                cell.number_format = '#,##0.00'
        net_cell = ws.cell(row=row_idx, column=12)
        net_cell.font = Font(bold=True, color='16A34A')
        row_idx += 1

    # Totals row
    totals = ['', 'TOTALS', '', '', '',
              float(run.total_gross or 0), '', '', '', '',
              float(run.total_deductions or 0), float(run.total_net or 0)]
    for ci, val in enumerate(totals, 1):
        cell = ws.cell(row=row_idx, column=ci, value=val)
        cell.fill = TOTAL_FILL
        cell.font = Font(bold=True, color='FFFFFF')
        cell.border = THIN
        if ci >= 6:
            cell.number_format = '#,##0.00'
    ws.cell(row=row_idx, column=12).fill = NET_FILL

    widths = [5, 16, 22, 18, 22, 14, 12, 12, 12, 12, 16, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = f'A4:L4'

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = '4:4'


def generate_payroll_excel(run, records, *, company_name='') -> bytes:
    wb = Workbook()

    # Sheet 1 — Summary
    ws_summary = wb.active
    ws_summary.title = 'Summary'
    _build_summary_sheet(ws_summary, run, records, company_name)

    # Sheet 2 — Payroll Detail
    ws_detail = wb.create_sheet(title='Payroll Detail')
    _build_detail_sheet(ws_detail, run, records, company_name)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
